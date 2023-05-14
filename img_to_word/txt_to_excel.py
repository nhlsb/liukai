# import openpyxl
#
# def convert_txt_to_excel(txt_file, excel_file):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#
#     with open(txt_file, 'r',encoding='utf-8') as file:
#         lines = file.readlines()
#         for row_num, line in enumerate(lines, start=1):
#             data = line.strip().split(' ')
#             if len(data) == 2:
#                 sheet.cell(row=row_num, column=1).value = data[0]
#                 sheet.cell(row=row_num, column=2).value = data[1]
#
#     wb.save(excel_file)
#     print("Conversion completed!")
#
# # 替换为你的TXT文件路径和Excel文件路径
# txt_file_path = 'a.txt'
# excel_file_path = 'excel_file.xlsx'
#
# convert_txt_to_excel(txt_file_path, excel_file_path)

import openpyxl

def convert_txt_to_excel(txt_file, excel_file):
    wb = openpyxl.Workbook()
    sheet = wb.active

    with open(txt_file, 'r',encoding='utf-8') as file:
        lines = file.readlines()
        for row_num, line in enumerate(lines, start=1):
            data = line.strip().split(' ')
            if len(data) == 3:
                sheet.cell(row=row_num, column=1).value = data[0]
                sheet.cell(row=row_num, column=2).value = data[1]
                sheet.cell(row=row_num, column=3).value = data[2]

    wb.save(excel_file)
    print("Conversion completed!")

# 替换为你的TXT文件路径和Excel文件路径
txt_file_path = 'a.txt'
excel_file_path = 'excel_file.xlsx'

convert_txt_to_excel(txt_file_path, excel_file_path)